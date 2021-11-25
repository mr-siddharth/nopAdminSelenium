import openpyxl
from configurations import testconfig
import traceback
import os.path
import shutil


def fetch_testdata(sheetname, td_file=testconfig.MASTER_TESTDATA_FILE,
                   testcase_name=None,
                   datastart_row=2):



    td_file = testconfig.TESTDATA_FILE = os.path.join("TestData", td_file)

    # Loading the test data excel file:
    wb = openpyxl.load_workbook(td_file)
    td_sheet = wb[sheetname]

    column_list = []
    # Reading the header row into a list:
    for i in range(1, td_sheet.max_column + 1):
        column_list.append(td_sheet.cell(row=1, column=i).value.strip())

    testcase_column_index = column_list.index("Test Case Name") + 1

    testdata = []
    # testdata will be a list of dictionaries where each dictionary will correspond to a row in test data sheet
    # which belong to the current test case.

    for i in range(2, td_sheet.max_row + 1):
        if (testcase_name is None) or (td_sheet.cell(row=i, column=testcase_column_index).value == testcase_name):
            # Converting the list of cell objects into a list of cell values through list comprehension
            # and then creating the dictionary by 'zip'ing the header row with the current data row:
            row_values = []
            for j in range(1, td_sheet.max_column + 1):
                row_values.append(td_sheet.cell(i, j).value)

            mydict = dict(zip(column_list, row_values))

            # Adding a rownum field to record the current row number
            mydict["rownum"] = i
            testdata.append(mydict)

    return testdata


def write_test_results_passed(sheetname, row_number):
    # copy the test data file if it doesn't already exist in the test run directory
    test_data_filename = os.path.split(testconfig.TESTDATA_FILE)[1]
    testconfig.TESTDATA_FILE = os.path.join(testconfig.TEST_RUN_DIR, test_data_filename)
    if not os.path.isfile(testconfig.TESTDATA_FILE):
        shutil.copyfile(os.path.join("TestData", test_data_filename), testconfig.TESTDATA_FILE)

    td_file = testconfig.TESTDATA_FILE
    td_wb = openpyxl.load_workbook(td_file)
    td_sheet = td_wb[sheetname]

    column_list = []
    # Reading the header row into a list:
    for i in range(1, td_sheet.max_column + 1):
        column_list = column_list + [td_sheet.cell(row=1, column=i).value.strip()]

    test_result_column_index = column_list.index("Test Result") + 1

    td_sheet.cell(row=row_number, column=test_result_column_index).value = "Passed"
    td_wb.save(td_file)


def write_test_results_failed(sheetname, row_number, error_message=""):
    # copy the test data file if it doesn't already exist in the test run directory
    test_data_filename = os.path.split(testconfig.TESTDATA_FILE)[1]
    testconfig.TESTDATA_FILE = os.path.join(testconfig.TEST_RUN_DIR, test_data_filename)
    if not os.path.isfile(testconfig.TESTDATA_FILE):
        shutil.copyfile(os.path.join("TestData", test_data_filename), testconfig.TESTDATA_FILE)

    td_file = testconfig.TESTDATA_FILE
    td_wb = openpyxl.load_workbook(td_file)
    td_sheet = td_wb[sheetname]

    column_list = []
    # Reading the header row into a list:
    for i in range(1, td_sheet.max_column + 1):
        column_list = column_list + [td_sheet.cell(row=1, column=i).value.strip()]

    test_result_column_index = column_list.index("Test Result") + 1
    errors_column_index = column_list.index("Errors If Any") + 1

    td_sheet.cell(row=row_number, column=test_result_column_index).value = "Failed"
    if error_message:
        td_sheet.cell(row=row_number, column=errors_column_index).value = error_message + '\n\n' + traceback.format_exc()
    else:
        td_sheet.cell(row=row_number, column=errors_column_index).value = traceback.format_exc()

    td_wb.save(td_file)